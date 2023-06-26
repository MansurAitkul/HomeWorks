import re

def extract_domains(email_list):
    domain_list = []
    for email in email_list:
        match = re.search(r'@(\w+\.\w+)', email)
        if match:
            domain = match.group(1)
            domain_list.append(domain)
    return domain_list

emails = [
    'john@example.com',
    'jane@test.com',
    'smith@gmail.com',
    'foo@bar.com'
]

domains = extract_domains(emails)
print(domains)



import re

def extract_vowel_words(text):
    vowel_words = re.findall(r'\b[aeiouAEIOU]\w+', text)
    return vowel_words

sentence = "An apple a day keeps the doctor away"
vowel_words = extract_vowel_words(sentence)
print(vowel_words)



import re

def split_string(text, delimiters):
    delimiter_pattern = '|'.join(map(re.escape, delimiters))
    splitted = re.split(delimiter_pattern, text)
    return splitted

string = "Hello, world! How are you today?"
delimiters = [',', ' ', '!']
splitted_string = split_string(string, delimiters)
print(splitted_string)


def count_votes(votes):
    vote_count = {}
    for candidate in votes:
        if candidate in vote_count:
            vote_count[candidate] += 1
        else:
            vote_count[candidate] = 1
    return vote_count

def determine_winner(vote_count):
    max_votes = max(vote_count.values())
    winners = [candidate for candidate, votes in vote_count.items() if votes == max_votes]
    if len(winners) == 1:
        return winners[0]
    else:
        winners.sort(key=lambda x: (len(x), x))
        return winners[0]

candidates = ["Аскаров", "Бекмуханов", "Ернур", "Пешая", "Карим", "Шаримазданов"]

votes = ["Ернур", "Аскаров", "Ернур", "Пешая", "Бекмуханов", "Ернур", "Шаримазданов"]

vote_count = count_votes(votes)
winner = determine_winner(vote_count)

print("Результаты голосования:")
for candidate, votes in vote_count.items():
    print(f"{candidate}: {votes} голосов")

print(f"\nПобедитель выборов: {winner}.")





import openpyxl


combined_wb = openpyxl.Workbook()
combined_ws = combined_wb.active


data_arrays = []
for i in range(1, 4):
    file_path = f'файл{i}.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data_array = []
    for row in ws.iter_rows(values_only=True):
        data_array.append(row)
    data_arrays.append(data_array)


for data_array in data_arrays:
    for row in data_array:
        combined_ws.append(row)


combined_wb.save('общий_файл.xlsx')

print("Данные успешно объединены и сохранены в новом файле.")